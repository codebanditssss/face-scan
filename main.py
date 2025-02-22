# # # import cv2
# # # import logging
# # # from datetime import datetime
# # # import os
# # # from src.face_processor.detector import FaceDetector
# # # from src.face_processor.matcher import FaceMatcher
# # # from src.database.student_db import StudentDatabase
# # # from src.database.attendance_db import AttendanceDatabase
# # # from src.utils.image_utils import ImageUtils
# # # from src.utils.excel_utils import ExcelManager
# # # from config import settings
# # # from config.thresholds import *


# # # def add_student_to_database(image_path, name, class_name=None, roll_number=None):
# # #     # Initialize face detector and database
# # #     face_detector = FaceDetector()
# # #     student_db = StudentDatabase()

# # #     # Load the image
# # #     image = cv2.imread(image_path)
# # #     if image is None:
# # #         print(f"Error: Unable to load image from {image_path}")
# # #         return

# # #     # Detect faces in the image
# # #     face_locations, face_encodings = face_detector.detect_faces(image)
# # #     if not face_encodings:
# # #         print("Error: No faces found in the image")
# # #         return

# # #     # Assuming only one face per image for simplicity
# # #     face_encoding = face_encodings[0]

# # #     # Add the student to the database
# # #     student_db.add_student(name, face_encoding.tobytes())
# # #     print(f"Added {name} to the database")

    
# # # class AttendanceSystem:
# # #     def __init__(self):
# # #         self.setup_logging()
# # #         self.initialize_components()
# # #         self.load_settings()

# # #     def setup_logging(self):
# # #         log_dir = settings.STORAGE['logs']
# # #         os.makedirs(log_dir, exist_ok=True)
# # #         log_file = os.path.join(log_dir, f'attendance_system_{datetime.now().strftime("%Y%m%d")}.log')
        
# # #         logging.basicConfig(
# # #             filename=log_file,
# # #             level=logging.INFO,
# # #             format=settings.LOGGING['format']
# # #         )
# # #         self.logger = logging.getLogger(__name__)

# # #     def initialize_components(self):
# # #         try:
# # #             self.face_detector = FaceDetector()
# # #             self.face_matcher = FaceMatcher()
# # #             self.student_db = StudentDatabase()
# # #             self.attendance_db = AttendanceDatabase()
# # #             self.image_utils = ImageUtils()
# # #             self.excel_manager = ExcelManager()
# # #             self.logger.info("All components initialized successfully")
# # #         except Exception as e:
# # #             self.logger.error(f"Error initializing components: {str(e)}")
# # #             raise

# # #     def load_settings(self):
# # #         self.camera_id = settings.CAMERA['device_id']
# # #         self.frame_width = settings.CAMERA['frame_width']
# # #         self.frame_height = settings.CAMERA['frame_height']
# # #         self.match_threshold = settings.FACE_RECOGNITION['match_threshold']

# # #     def process_frame(self, frame):
# # #         try:
# # #             # Detect faces
# # #             face_locations, processed_faces = self.face_detector.detect_and_process(frame)
            
# # #             for idx, (face_location, face_img) in enumerate(zip(face_locations, processed_faces)):
# # #                 # Find best match
# # #                 student_id, match_score, student_info = self.face_matcher.find_best_match(
# # #                     face_img, self.match_threshold)
                
# # #                 if student_id and student_info:
# # #                     # Mark attendance
# # #                     attendance_data = {
# # #                         'student_id': student_id,
# # #                         'date': datetime.now().date(),
# # #                         'entry_time': datetime.now().time(),
# # #                         'match_confidence': match_score
# # #                     }
# # #                     self.attendance_db.mark_attendance(attendance_data)
                    
# # #                     # Draw on frame
# # #                     frame = self.image_utils.draw_face_box(
# # #                         frame, face_location,
# # #                         name=student_info.get('name', ''),
# # #                         confidence=match_score
# # #                     )
# # #                 else:
# # #                     frame = self.image_utils.draw_face_box(
# # #                         frame, face_location,
# # #                         name="Unknown",
# # #                         confidence=match_score if match_score else 0.0,
# # #                         color=(0, 0, 255)
# # #                     )
            
# # #             return frame
            
# # #         except Exception as e:
# # #             self.logger.error(f"Error processing frame: {str(e)}")
# # #             return frame

# # #     def run(self):
# # #         try:
# # #             cap = cv2.VideoCapture(self.camera_id)
# # #             cap.set(cv2.CAP_PROP_FRAME_WIDTH, self.frame_width)
# # #             cap.set(cv2.CAP_PROP_FRAME_HEIGHT, self.frame_height)
            
# # #             self.logger.info("Starting attendance system")
            
# # #             while True:
# # #                 ret, frame = cap.read()
# # #                 if not ret:
# # #                     break
                
# # #                 # Process frame
# # #                 processed_frame = self.process_frame(frame)
                
# # #                 # Display
# # #                 cv2.imshow(settings.UI['window_title'], processed_frame)
                
# # #                 # Break on 'q'
# # #                 if cv2.waitKey(1) & 0xFF == ord('q'):
# # #                     break
                    
# # #             cap.release()
# # #             cv2.destroyAllWindows()
            
# # #         except Exception as e:
# # #             self.logger.error(f"Error in main loop: {str(e)}")
# # #         finally:
# # #             self.cleanup()

# # #     def cleanup(self):
# # #         try:
# # #             # Generate daily reports
# # #             self.generate_reports()
            
# # #             # Backup database
# # #             if settings.BACKUP['auto_backup']:
# # #                 self.backup_database()
                
# # #             self.logger.info("System shutdown complete")
            
# # #         except Exception as e:
# # #             self.logger.error(f"Error during cleanup: {str(e)}")

# # #     def generate_reports(self):
# # #         try:
# # #             today = datetime.now().date()
            
# # #             # Get all classes
# # #             classes = self.student_db.get_class_list()
            
# # #             for class_name in classes:
# # #                 # Get attendance data
# # #                 attendance_data = self.attendance_db.get_daily_attendance(
# # #                     class_name, today)
                
# # #                 # Generate Excel report
# # #                 self.excel_manager.create_daily_report(
# # #                     attendance_data, class_name, today)
                
# # #         except Exception as e:
# # #             self.logger.error(f"Error generating reports: {str(e)}")

# # #     def backup_database(self):
# # #         try:
# # #             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
# # #             backup_path = os.path.join(
# # #                 settings.DATABASE['backup_dir'],
# # #                 f'attendance_backup_{timestamp}.db'
# # #             )
            
# # #             # Implement database backup logic here
# # #             self.logger.info(f"Database backed up to: {backup_path}")
            
# # #         except Exception as e:
# # #             self.logger.error(f"Error backing up database: {str(e)}")

# # # if __name__ == "__main__":
# # #     add_student_to_database(
# # #         image_path='data/student_images/khushi diwan.jpg',  # Update this path to the actual image file
# # #         name='Khushi Diwan',
# # #         class_name='12th A',  # Optional
# # #         roll_number='2481069'   # Optional
# # #     )
# # #     system = AttendanceSystem()
# # #     system.run()

# # import cv2
# # import logging
# # from datetime import datetime
# # import os
# # from src.face_processor.detector import FaceDetector
# # from src.face_processor.matcher import FaceMatcher
# # from src.database.student_db import StudentDatabase
# # from src.database.attendance_db import AttendanceDatabase
# # from src.utils.image_utils import ImageUtils
# # from src.utils.excel_utils import ExcelManager
# # from config import settings
# # from config.thresholds import *

# # def add_student_to_database(image_path, name, class_name=None, roll_number=None):
# #     """
# #     Add a student's face to the face recognition database.
    
# #     Args:
# #         image_path (str): Path to the student's image
# #         name (str): Student's full name
# #         class_name (str, optional): Student's class
# #         roll_number (str, optional): Student's roll number
# #     """
# #     # Initialize components
# #     face_detector = FaceDetector()
# #     face_matcher = FaceMatcher()
# #     student_db = StudentDatabase()

# #     # Load the image
# #     image = cv2.imread(image_path)
# #     if image is None:
# #         print(f"Error: Unable to load image from {image_path}")
# #         return None

# #     # Detect and process faces
# #     face_locations, processed_faces = face_detector.detect_and_process(image)
    
# #     if not processed_faces:
# #         print("Error: No faces found in the image")
# #         return None

# #     # Take the first detected face
# #     processed_face = processed_faces[0]

# #     # Prepare student data
# #     student_id = name.replace(' ', '_').lower()
# #     metadata = {
# #         'name': name,
# #         'class': class_name,
# #         'roll_number': roll_number
# #     }
    
# #     # Store face in matcher database
# #     face_matcher.add_face(
# #         student_id=student_id, 
# #         face_image=processed_face, 
# #         metadata=metadata
# #     )
    
# #     # Store in student database
# #     student_db.add_student(
# #         name=name, 
# #         face_encoding=processed_face.tobytes(),
# #         class_name=class_name,
# #         roll_number=roll_number
# #     )
    
# #     print(f"Added {name} to the face and student databases")
# #     return student_id

# # class AttendanceSystem:
# #     def __init__(self):
# #         self.setup_logging()
# #         self.initialize_components()
# #         self.load_settings()

# #     def setup_logging(self):
# #         log_dir = settings.STORAGE['logs']
# #         os.makedirs(log_dir, exist_ok=True)
# #         log_file = os.path.join(log_dir, f'attendance_system_{datetime.now().strftime("%Y%m%d")}.log')
        
# #         logging.basicConfig(
# #             filename=log_file,
# #             level=logging.INFO,
# #             format=settings.LOGGING['format']
# #         )
# #         self.logger = logging.getLogger(__name__)

# #     def initialize_components(self):
# #         try:
# #             self.face_detector = FaceDetector()
# #             self.face_matcher = FaceMatcher()
# #             self.student_db = StudentDatabase()
# #             self.attendance_db = AttendanceDatabase()
# #             self.image_utils = ImageUtils()
# #             self.excel_manager = ExcelManager()
# #             self.logger.info("All components initialized successfully")
# #         except Exception as e:
# #             self.logger.error(f"Error initializing components: {str(e)}")
# #             raise

# #     def load_settings(self):
# #         self.camera_id = settings.CAMERA['device_id']
# #         self.frame_width = settings.CAMERA['frame_width']
# #         self.frame_height = settings.CAMERA['frame_height']
# #         self.match_threshold = settings.FACE_RECOGNITION['match_threshold']

# #     def process_frame(self, frame):
# #         try:
# #             # Detect faces
# #             face_locations, processed_faces = self.face_detector.detect_and_process(frame)
            
# #             for idx, (face_location, face_img) in enumerate(zip(face_locations, processed_faces)):
# #                 # Find best match
# #                 student_id, match_score, student_info = self.face_matcher.find_best_match(
# #                     face_img, self.match_threshold)
                
# #                 if student_id and student_info:
# #                     # Mark attendance
# #                     attendance_data = {
# #                         'student_id': student_id,
# #                         'date': datetime.now().date(),
# #                         'entry_time': datetime.now().time(),
# #                         'match_confidence': match_score
# #                     }
# #                     self.attendance_db.mark_attendance(attendance_data)
                    
# #                     # Draw on frame
# #                     frame = self.image_utils.draw_face_box(
# #                         frame, face_location,
# #                         name=student_info.get('name', ''),
# #                         confidence=match_score
# #                     )
# #                 else:
# #                     frame = self.image_utils.draw_face_box(
# #                         frame, face_location,
# #                         name="Unknown",
# #                         confidence=match_score if match_score else 0.0,
# #                         color=(0, 0, 255)
# #                     )
            
# #             return frame
            
# #         except Exception as e:
# #             self.logger.error(f"Error processing frame: {str(e)}")
# #             return frame

# #     def run(self):
# #         try:
# #             cap = cv2.VideoCapture(self.camera_id)
# #             cap.set(cv2.CAP_PROP_FRAME_WIDTH, self.frame_width)
# #             cap.set(cv2.CAP_PROP_FRAME_HEIGHT, self.frame_height)
            
# #             self.logger.info("Starting attendance system")
            
# #             while True:
# #                 ret, frame = cap.read()
# #                 if not ret:
# #                     break
                
# #                 # Process frame
# #                 processed_frame = self.process_frame(frame)
                
# #                 # Display
# #                 cv2.imshow(settings.UI['window_title'], processed_frame)
                
# #                 # Break on 'q'
# #                 if cv2.waitKey(1) & 0xFF == ord('q'):
# #                     break
                    
# #             cap.release()
# #             cv2.destroyAllWindows()
            
# #         except Exception as e:
# #             self.logger.error(f"Error in main loop: {str(e)}")
# #         finally:
# #             self.cleanup()

# #     def cleanup(self):
# #         try:
# #             # Generate daily reports
# #             self.generate_reports()
            
# #             # Backup database
# #             if settings.BACKUP['auto_backup']:
# #                 self.backup_database()
                
# #             self.logger.info("System shutdown complete")
            
# #         except Exception as e:
# #             self.logger.error(f"Error during cleanup: {str(e)}")

# #     def generate_reports(self):
# #         try:
# #             today = datetime.now().date()
            
# #             # Get all classes
# #             classes = self.student_db.get_class_list()
            
# #             for class_name in classes:
# #                 # Get attendance data
# #                 attendance_data = self.attendance_db.get_daily_attendance(
# #                     class_name, today)
                
# #                 # Generate Excel report
# #                 self.excel_manager.create_daily_report(
# #                     attendance_data, class_name, today)
                
# #         except Exception as e:
# #             self.logger.error(f"Error generating reports: {str(e)}")

# #     def backup_database(self):
# #         try:
# #             timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
# #             backup_path = os.path.join(
# #                 settings.DATABASE['backup_dir'],
# #                 f'attendance_backup_{timestamp}.db'
# #             )
            
# #             # Implement database backup logic here
# #             self.logger.info(f"Database backed up to: {backup_path}")
            
# #         except Exception as e:
# #             self.logger.error(f"Error backing up database: {str(e)}")

# # if __name__ == "__main__":
# #     # Add your student to the database before running the system
# #     add_student_to_database(
# #         image_path='data/student_images/khushi diwan.jpg',
# #         name='Khushi Diwan',
# #         class_name='12th A',
# #         roll_number='2481069'
# #     )
    
# #     # Run the attendance system
# #     system = AttendanceSystem()
# #     system.run()


# import cv2
# import logging
# from datetime import datetime
# import os
# import sys
# from openpyxl import Workbook

# from src.face_processor.detector import FaceDetector
# from src.face_processor.matcher import FaceMatcher
# from src.database.student_db import StudentDatabase
# from src.database.attendance_db import AttendanceDatabase
# from src.utils.image_utils import ImageUtils
# from src.utils.excel_utils import ExcelManager
# from config import settings
# from config.thresholds import *

# def create_attendance_excel(student_info, match_score):
#     """
#     Create an Excel file with student attendance details.
    
#     Args:
#         student_info (dict): Dictionary containing student information
#         match_score (float): Face match confidence score
    
#     Returns:
#         str: Path to the created Excel file
#     """
#     # Ensure directory exists
#     os.makedirs('attendance_records', exist_ok=True)
    
#     # Create filename with current date
#     filename = f'attendance_records/attendance_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
#     # Create workbook and select active sheet
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Attendance Record"
    
#     # Add headers
#     headers = [
#         'Roll No', 'Name', 'Entry Time', 'Exit Time', 
#         'Status', 'Match Confidence', 'Remarks'
#     ]
#     for col, header in enumerate(headers, start=1):
#         ws.cell(row=1, column=col, value=header)
    
#     # Add student data
#     data = [
#         student_info.get('roll_number', 'N/A'),  # Roll No
#         student_info.get('name', 'Unknown'),     # Name
#         datetime.now().strftime("%H:%M:%S"),     # Entry Time
#         '',                                      # Exit Time (blank)
#         'Present',                               # Status
#         f"{match_score:.2f}%",                   # Match Confidence
#         'Attendance Marked'                      # Remarks
#     ]
    
#     for col, value in enumerate(data, start=1):
#         ws.cell(row=2, column=col, value=value)
    
#     # Save the workbook
#     wb.save(filename)
    
#     print(f"Attendance record saved to {filename}")
#     return filename

# def add_student_to_database(image_path, name, class_name=None, roll_number=None):
#     """
#     Add a student to the face recognition database.
    
#     Args:
#         image_path (str): Path to student's image
#         name (str): Student's full name
#         class_name (str, optional): Student's class
#         roll_number (str, optional): Student's roll number
    
#     Returns:
#         str or None: Student ID if successful, None otherwise
#     """
#     # Initialize components
#     face_detector = FaceDetector()
#     face_matcher = FaceMatcher()
#     student_db = StudentDatabase()

#     # Load the image
#     image = cv2.imread(image_path)
#     if image is None:
#         print(f"Error: Unable to load image from {image_path}")
#         return None

#     # Detect and process faces
#     face_locations, processed_faces = face_detector.detect_and_process(image)
    
#     if not processed_faces:
#         print("Error: No faces found in the image")
#         return None

#     # Take the first detected face
#     processed_face = processed_faces[0]

#     # Prepare student data
#     student_id = name.replace(' ', '_').lower()
#     metadata = {
#         'name': name,
#         'class': class_name,
#         'roll_number': roll_number
#     }
    
#     # Store face in matcher database
#     face_matcher.add_face(
#         student_id=student_id, 
#         face_image=processed_face, 
#         metadata=metadata
#     )
    
#     # Store in student database
#     student_db.add_student(
#         name=name, 
#         face_encoding=processed_face.tobytes(),
#         class_name=class_name,
#         roll_number=roll_number
#     )
    
#     # print(f"Added {name} to the face and student databases")
#     return student_id

# class AttendanceSystem:
#     def __init__(self):
#         self.setup_logging()
#         self.initialize_components()
#         self.load_settings()
#         self.processed_students = set()

#     def setup_logging(self):
#         """Setup logging configuration."""
#         log_dir = settings.STORAGE['logs']
#         os.makedirs(log_dir, exist_ok=True)
#         log_file = os.path.join(log_dir, f'attendance_system_{datetime.now().strftime("%Y%m%d")}.log')
        
#         logging.basicConfig(
#             filename=log_file,
#             level=logging.INFO,
#             format=settings.LOGGING['format']
#         )
#         self.logger = logging.getLogger(__name__)

#     def initialize_components(self):
#         """Initialize system components."""
#         try:
#             self.face_detector = FaceDetector()
#             self.face_matcher = FaceMatcher()
#             self.student_db = StudentDatabase()
#             self.attendance_db = AttendanceDatabase()
#             self.image_utils = ImageUtils()
#             self.excel_manager = ExcelManager()
#             self.logger.info("All components initialized successfully")
#         except Exception as e:
#             self.logger.error(f"Error initializing components: {str(e)}")
#             raise

#     def load_settings(self):
#         """Load camera and recognition settings."""
#         self.camera_id = settings.CAMERA['device_id']
#         self.frame_width = settings.CAMERA['frame_width']
#         self.frame_height = settings.CAMERA['frame_height']
#         self.match_threshold = settings.FACE_RECOGNITION['match_threshold']

#     def process_frame(self, frame):
#         """
#         Process a single video frame for face recognition.
        
#         Args:
#             frame (numpy.ndarray): Input video frame
        
#         Returns:
#             numpy.ndarray: Processed frame with annotations
#         """
#         try:
#             # Detect faces
#             face_locations, processed_faces = self.face_detector.detect_and_process(frame)
            
#             for idx, (face_location, face_img) in enumerate(zip(face_locations, processed_faces)):
#                 # Find best match
#                 student_id, match_score, student_info = self.face_matcher.find_best_match(
#                     face_img, self.match_threshold)
                
#                 # Check if student already processed
#                 if student_id and student_info and match_score >= 70:
#                     if student_id not in self.processed_students:
#                         # Create Excel attendance record
#                         excel_file = create_attendance_excel(student_info, match_score)
                        
#                         # Mark student as processed
#                         self.processed_students.add(student_id)
                        
#                         # Optional attendance database marking
#                         attendance_data = {
#                             'student_id': student_id,
#                             'date': datetime.now().date(),
#                             'entry_time': datetime.now().time(),
#                             'match_confidence': match_score
#                         }
#                         self.attendance_db.mark_attendance(attendance_data)
                        
#                         # Draw on frame
#                         frame = self.image_utils.draw_face_box(
#                             frame, face_location,
#                             name=f"{student_info.get('name', '')} - Attendance Recorded",
#                             confidence=match_score
#                         )
                    
#                     # Stop after first match
#                     break
#                 else:
#                     frame = self.image_utils.draw_face_box(
#                         frame, face_location,
#                         name="Unknown",
#                         confidence=match_score if match_score else 0.0,
#                         color=(0, 0, 255)
#                     )
            
#             return frame
            
#         except Exception as e:
#             self.logger.error(f"Error processing frame: {str(e)}")
#             return frame

#     def run(self):
#         """
#         Run the attendance system using camera input.
#         """
#         try:
#             cap = cv2.VideoCapture(self.camera_id)
#             cap.set(cv2.CAP_PROP_FRAME_WIDTH, self.frame_width)
#             cap.set(cv2.CAP_PROP_FRAME_HEIGHT, self.frame_height)
            
#             self.logger.info("Starting attendance system")
            
#             while True:
#                 ret, frame = cap.read()
#                 if not ret:
#                     break
                
#                 # Process frame
#                 processed_frame = self.process_frame(frame)
                
#                 # Display
#                 cv2.imshow(settings.UI['window_title'], processed_frame)
                
#                 # Break on 'q' or if attendance is recorded
#                 if cv2.waitKey(1) & 0xFF == ord('q') or len(self.processed_students) > 0:
#                     break
                    
#             cap.release()
#             cv2.destroyAllWindows()
            
#         except Exception as e:
#             self.logger.error(f"Error in main loop: {str(e)}")
#         finally:
#             self.cleanup()

#     def cleanup(self):
#         """Perform system cleanup operations."""
#         try:
#             # Generate final reports if needed
#             self.logger.info("System shutdown complete")
#         except Exception as e:
#             self.logger.error(f"Error during cleanup: {str(e)}")

# if __name__ == "__main__":
#     # Add student to database before running system
#     add_student_to_database(
#         image_path='data/student_images/khushi diwan.jpg',
#         name='Khushi Diwan',
#         class_name='12th A',
#         roll_number='2481069'
#     )
    
#     # Run the attendance system
#     system = AttendanceSystem()
#     system.run()

# import cv2
# import logging
# from datetime import datetime
# import os
# import sys
# from openpyxl import load_workbook, Workbook

# from src.face_processor.detector import FaceDetector
# from src.face_processor.matcher import FaceMatcher
# from src.database.student_db import StudentDatabase
# from src.database.attendance_db import AttendanceDatabase
# from src.utils.image_utils import ImageUtils
# from src.utils.excel_utils import ExcelManager
# from config import settings
# from config.thresholds import *

# def create_or_update_attendance_excel(student_info, match_score):
#     """
#     Create or update an attendance Excel file.
    
#     Args:
#         student_info (dict): Student details
#         match_score (float): Face match confidence
    
#     Returns:
#         str: Path to the Excel file
#     """
#     # Ensure directory exists
#     os.makedirs('attendance_records', exist_ok=True)
#     filename = 'attendance_records/daily_attendance.xlsx'
    
#     # Check if file exists
#     if not os.path.exists(filename):
#         # Create new workbook
#         wb = Workbook()
#         ws = wb.active
#         ws.title = "Attendance Record"
#         headers = [
#             'Roll No', 'Name', 'Entry Time', 'Exit Time', 
#             'Status', 'Match Confidence', 'Remarks'
#         ]
#         for col, header in enumerate(headers, start=1):
#             ws.cell(row=1, column=col, value=header)
#     else:
#         # Load existing workbook
#         wb = load_workbook(filename)
#         ws = wb.active
    
#     # Check if student already exists
#     existing_rows = list(ws.rows)
#     student_exists = any(
#         row[0].value == student_info.get('roll_number') 
#         for row in existing_rows[1:]
#     )
    
#     if not existing_rows:
#         # This handles unlikely edge case of empty file
#         return filename
    
#     if not existing_rows[1:]:
#         max_row = 1
#     else:
#         max_row = len(existing_rows)
    
#     # Add new student if not exists
#     if not existing_rows or not student_exists:
#         data = [
#             student_info.get('roll_number', 'N/A'),
#             student_info.get('name', 'Unknown'),
#             datetime.now().strftime("%H:%M:%S"),
#             '',
#             'Present',
#             f"{match_score:.2f}%",
#             'Attendance Marked'
#         ]
        
#         for col, value in enumerate(data, start=1):
#             ws.cell(row=max_row + 1, column=col, value=value)
    
#     # Save the workbook
#     wb.save(filename)
    
#     print(f"Attendance record updated in {filename}")
#     return filename

# def add_student_to_database(image_path, name, class_name=None, roll_number=None):
#     """
#     Add a student to the face recognition database.
    
#     Args:
#         image_path (str): Path to student's image
#         name (str): Student's full name
#         class_name (str, optional): Student's class
#         roll_number (str, optional): Student's roll number
    
#     Returns:
#         str or None: Student ID if successful, None otherwise
#     """
#     # Initialize components
#     face_detector = FaceDetector()
#     face_matcher = FaceMatcher()
#     student_db = StudentDatabase()

#     # Load the image
#     image = cv2.imread(image_path)
#     if image is None:
#         print(f"Error: Unable to load image from {image_path}")
#         return None

#     # Detect and process faces
#     face_locations, processed_faces = face_detector.detect_and_process(image)
    
#     if not processed_faces:
#         print("Error: No faces found in the image")
#         return None

#     # Take the first detected face
#     processed_face = processed_faces[0]

#     # Prepare student data
#     student_id = name.replace(' ', '_').lower()
#     metadata = {
#         'name': name,
#         'class': class_name,
#         'roll_number': roll_number
#     }
    
#     # Store face in matcher database
#     face_matcher.add_face(
#         student_id=student_id, 
#         face_image=processed_face, 
#         metadata=metadata
#     )
    
#     # Store in student database
#     student_db.add_student(
#         name=name, 
#         face_encoding=processed_face.tobytes(),
#         class_name=class_name,
#         roll_number=roll_number
#     )
    
#     # print(f"Added {name} to the face and student databases")
#     return student_id

# def add_multiple_students():
#     """
#     Add multiple students to the database.
#     """
#     students = [
#         {
#             'name': 'Khushi Diwan',
#             'image_path': 'data/student_images/khushi diwan.jpg',
#             'class_name': '12th A',
#             'roll_number': '2481069'
#         },
#         {
#             'name': 'Devansh Datta',
#             'image_path': 'data/student_images/devansh datta.jpg',
#             'class_name': '12th A',
#             'roll_number': '2481436'
#         },
#         {
#             'name': 'Pooja Batra',
#             'image_path': 'data/student_images/pooja batra.jpg',
#             'class_name': '12th A',
#             'roll_number': '2480236'
#         },
#         {
#             'name': 'Aarti Chugh',
#             'image_path': 'data/student_images/aarti chugh.jpg',
#             'class_name': '12th A',
#             'roll_number': '2580236'
#         }
#     ]
    
#     for student in students:
#         add_student_to_database(
#             image_path=student['image_path'],
#             name=student['name'],
#             class_name=student['class_name'],
#             roll_number=student['roll_number']
#         )

# class AttendanceSystem:
#     def __init__(self):
#         self.setup_logging()
#         self.initialize_components()
#         self.load_settings()
#         self.processed_students = set()

#     def setup_logging(self):
#         """Setup logging configuration."""
#         log_dir = settings.STORAGE['logs']
#         os.makedirs(log_dir, exist_ok=True)
#         log_file = os.path.join(log_dir, f'attendance_system_{datetime.now().strftime("%Y%m%d")}.log')
        
#         logging.basicConfig(
#             filename=log_file,
#             level=logging.INFO,
#             format=settings.LOGGING['format']
#         )
#         self.logger = logging.getLogger(__name__)

#     def initialize_components(self):
#         """Initialize system components."""
#         try:
#             self.face_detector = FaceDetector()
#             self.face_matcher = FaceMatcher()
#             self.student_db = StudentDatabase()
#             self.attendance_db = AttendanceDatabase()
#             self.image_utils = ImageUtils()
#             self.excel_manager = ExcelManager()
#             self.logger.info("All components initialized successfully")
#         except Exception as e:
#             self.logger.error(f"Error initializing components: {str(e)}")
#             raise

#     def load_settings(self):
#         """Load camera and recognition settings."""
#         self.camera_id = settings.CAMERA['device_id']
#         self.frame_width = settings.CAMERA['frame_width']
#         self.frame_height = settings.CAMERA['frame_height']
#         self.match_threshold = settings.FACE_RECOGNITION['match_threshold']

#     def process_frame(self, frame):
#         """
#         Process a single video frame for face recognition.
        
#         Args:
#             frame (numpy.ndarray): Input video frame
        
#         Returns:
#             numpy.ndarray: Processed frame with annotations
#         """
#         try:
#             # Detect faces
#             face_locations, processed_faces = self.face_detector.detect_and_process(frame)
            
#             for idx, (face_location, face_img) in enumerate(zip(face_locations, processed_faces)):
#                 # Find best match with lower threshold
#                 student_id, match_score, student_info = self.face_matcher.find_best_match(
#                     face_img, min_threshold=60)
                
#                 # # Print debug information
#                 # print(f"Detection: ID={student_id}, Score={match_score}, Info={student_info}")
                
#                 # Check if student already processed
#                 if student_id and student_info and match_score >= 30:
#                     if student_id not in self.processed_students:
#                         # Create/Update Excel attendance record
#                         excel_file = create_or_update_attendance_excel(student_info, match_score)
                        
#                         # Mark student as processed
#                         self.processed_students.add(student_id)
                        
#                         # Optional attendance database marking
#                         attendance_data = {
#                             'student_id': student_id,
#                             'date': datetime.now().date(),
#                             'entry_time': datetime.now().time(),
#                             'match_confidence': match_score
#                         }
#                         self.attendance_db.mark_attendance(attendance_data)
                        
#                         # Draw on frame
#                         frame = self.image_utils.draw_face_box(
#                             frame, face_location,
#                             name=f"{student_info.get('name', '')} - Attendance Recorded",
#                             confidence=match_score
#                         )
                    
#                     # Stop after first match
#                     break
#                 else:
#                     frame = self.image_utils.draw_face_box(
#                         frame, face_location,
#                         name="Unknown",
#                         confidence=match_score if match_score else 0.0,
#                         color=(0, 0, 255)
#                     )
            
#             return frame
            
#         except Exception as e:
#             self.logger.error(f"Error processing frame: {str(e)}")
#             return frame

#     def run(self):
#         """
#         Run the attendance system using camera input.
#         """
#         try:
#             cap = cv2.VideoCapture(self.camera_id)
#             cap.set(cv2.CAP_PROP_FRAME_WIDTH, self.frame_width)
#             cap.set(cv2.CAP_PROP_FRAME_HEIGHT, self.frame_height)
            
#             self.logger.info("Starting attendance system")
            
#             while True:
#                 ret, frame = cap.read()
#                 if not ret:
#                     break
                
#                 # Process frame
#                 processed_frame = self.process_frame(frame)
                
#                 # Display
#                 cv2.imshow(settings.UI['window_title'], processed_frame)
                
#                 # Break on 'q' or if attendance is recorded
#                 if cv2.waitKey(1) & 0xFF == ord('q') or len(self.processed_students) > 0:
#                     break
                    
#             cap.release()
#             cv2.destroyAllWindows()
            
#         except Exception as e:
#             self.logger.error(f"Error in main loop: {str(e)}")
#         finally:
#             self.cleanup()

#     def cleanup(self):
#         """Perform system cleanup operations."""
#         try:
#             # Generate final reports if needed
#             self.logger.info("System shutdown complete")
#         except Exception as e:
#             self.logger.error(f"Error during cleanup: {str(e)}")

# if __name__ == "__main__":
#     # Add multiple students to database
#     add_multiple_students()
    
#     # Run the attendance system
#     system = AttendanceSystem()
#     system.run()


import cv2
import logging
from datetime import datetime
import os
import sys
from openpyxl import load_workbook, Workbook
import numpy as np
import base64
import asyncio
import uvicorn
from fastapi import FastAPI, WebSocket, HTTPException
from fastapi.middleware.cors import CORSMiddleware

from src.face_processor.detector import FaceDetector
from src.face_processor.matcher import FaceMatcher
from src.database.student_db import StudentDatabase
from src.database.attendance_db import AttendanceDatabase
from src.utils.image_utils import ImageUtils
from src.utils.excel_utils import ExcelManager
from config import settings
from config.thresholds import *

# Initialize FastAPI
app = FastAPI()

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],  # In production, replace with your frontend URL
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def create_or_update_attendance_excel(student_info, match_score):
    """
    Create or update an attendance Excel file.
    """
    # Ensure directory exists
    os.makedirs('attendance_records', exist_ok=True)
    filename = 'attendance_records/daily_attendance.xlsx'
    
    # Check if file exists
    if not os.path.exists(filename):
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendance Record"
        headers = [
            'Roll No', 'Name', 'Entry Time', 'Exit Time', 
            'Status', 'Match Confidence', 'Remarks'
        ]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
    else:
        # Load existing workbook
        wb = load_workbook(filename)
        ws = wb.active
    
    # Check if student already exists
    existing_rows = list(ws.rows)
    student_exists = any(
        row[0].value == student_info.get('roll_number') 
        for row in existing_rows[1:]
    )
    
    if not existing_rows:
        # This handles unlikely edge case of empty file
        return filename
    
    if not existing_rows[1:]:
        max_row = 1
    else:
        max_row = len(existing_rows)
    
    # Add new student if not exists
    if not existing_rows or not student_exists:
        data = [
            student_info.get('roll_number', 'N/A'),
            student_info.get('name', 'Unknown'),
            datetime.now().strftime("%H:%M:%S"),
            '',
            'Present',
            f"{match_score:.2f}%",
            'Attendance Marked'
        ]
        
        for col, value in enumerate(data, start=1):
            ws.cell(row=max_row + 1, column=col, value=value)
    
    # Save the workbook
    wb.save(filename)
    
    print(f"Attendance record updated in {filename}")
    return filename

def add_student_to_database(image_path, name, class_name=None, roll_number=None):
    """
    Add a student to the face recognition database.
    """
    # Initialize components
    face_detector = FaceDetector()
    face_matcher = FaceMatcher()
    student_db = StudentDatabase()

    # Load the image
    image = cv2.imread(image_path)
    if image is None:
        print(f"Error: Unable to load image from {image_path}")
        return None

    # Detect and process faces
    face_locations, processed_faces = face_detector.detect_and_process(image)
    
    if not processed_faces:
        print("Error: No faces found in the image")
        return None

    # Take the first detected face
    processed_face = processed_faces[0]

    # Prepare student data
    student_id = name.replace(' ', '_').lower()
    metadata = {
        'name': name,
        'class': class_name,
        'roll_number': roll_number
    }
    
    # Store face in matcher database
    face_matcher.add_face(
        student_id=student_id, 
        face_image=processed_face, 
        metadata=metadata
    )
    
    # Store in student database
    student_db.add_student(
        name=name, 
        face_encoding=processed_face.tobytes(),
        class_name=class_name,
        roll_number=roll_number
    )
    
    return student_id

def add_multiple_students():
    """
    Add multiple students to the database.
    """
    students = [
        {
            'name': 'Khushi Diwan',
            'image_path': 'data/student_images/khushi diwan.jpg',
            'class_name': '12th A',
            'roll_number': '2481069'
        },
        {
            'name': 'Devansh Datta',
            'image_path': 'data/student_images/devansh datta.jpg',
            'class_name': '12th A',
            'roll_number': '2481436'
        },
        {
            'name': 'Pooja Batra',
            'image_path': 'data/student_images/pooja batra.jpg',
            'class_name': '12th A',
            'roll_number': '2480236'
        },
        {
            'name': 'Aarti Chugh',
            'image_path': 'data/student_images/aarti chugh.jpg',
            'class_name': '12th A',
            'roll_number': '2580236'
        }
    ]
    
    for student in students:
        add_student_to_database(
            image_path=student['image_path'],
            name=student['name'],
            class_name=student['class_name'],
            roll_number=student['roll_number']
        )

class AttendanceSystem:
    def __init__(self):
        self.setup_logging()
        self.initialize_components()
        self.load_settings()
        self.processed_students = set()

    def setup_logging(self):
        """Setup logging configuration."""
        log_dir = settings.STORAGE['logs']
        os.makedirs(log_dir, exist_ok=True)
        log_file = os.path.join(log_dir, f'attendance_system_{datetime.now().strftime("%Y%m%d")}.log')
        
        logging.basicConfig(
            filename=log_file,
            level=logging.INFO,
            format=settings.LOGGING['format']
        )
        self.logger = logging.getLogger(__name__)

    def initialize_components(self):
        """Initialize system components."""
        try:
            self.face_detector = FaceDetector()
            self.face_matcher = FaceMatcher()
            self.student_db = StudentDatabase()
            self.attendance_db = AttendanceDatabase()
            self.image_utils = ImageUtils()
            self.excel_manager = ExcelManager()
            self.logger.info("All components initialized successfully")
        except Exception as e:
            self.logger.error(f"Error initializing components: {str(e)}")
            raise

    def load_settings(self):
        """Load camera and recognition settings."""
        self.camera_id = settings.CAMERA['device_id']
        self.frame_width = settings.CAMERA['frame_width']
        self.frame_height = settings.CAMERA['frame_height']
        self.match_threshold = settings.FACE_RECOGNITION['match_threshold']

    def process_frame(self, frame):
        """
        Process a single video frame for face recognition.
        """
        try:
            # Detect faces
            face_locations, processed_faces = self.face_detector.detect_and_process(frame)
            
            for idx, (face_location, face_img) in enumerate(zip(face_locations, processed_faces)):
                # Find best match with lower threshold
                student_id, match_score, student_info = self.face_matcher.find_best_match(
                    face_img, min_threshold=60)
                
                # Check if student already processed
                if student_id and student_info and match_score >= 30:
                    if student_id not in self.processed_students:
                        # Create/Update Excel attendance record
                        excel_file = create_or_update_attendance_excel(student_info, match_score)
                        
                        # Mark student as processed
                        self.processed_students.add(student_id)
                        
                        # Optional attendance database marking
                        attendance_data = {
                            'student_id': student_id,
                            'date': datetime.now().date(),
                            'entry_time': datetime.now().time(),
                            'match_confidence': match_score
                        }
                        self.attendance_db.mark_attendance(attendance_data)
                        
                        # Draw on frame
                        frame = self.image_utils.draw_face_box(
                            frame, face_location,
                            name=f"{student_info.get('name', '')} - Attendance Recorded",
                            confidence=match_score
                        )
                    
                    # Stop after first match
                    break
                else:
                    frame = self.image_utils.draw_face_box(
                        frame, face_location,
                        name="Unknown",
                        confidence=match_score if match_score else 0.0,
                        color=(0, 0, 255)
                    )
            
            return frame
            
        except Exception as e:
            self.logger.error(f"Error processing frame: {str(e)}")
            return frame

    async def process_frame_websocket(self, frame):
        """Process a frame for websocket connection."""
        try:
            # Make a copy for drawing
            display_frame = frame.copy()
            
            # Detect faces
            face_locations, processed_faces = self.face_detector.detect_and_process(frame)
            
            if not face_locations or not processed_faces:
                # Convert frame to base64
                _, buffer = cv2.imencode('.jpg', display_frame)
                frame_b64 = base64.b64encode(buffer).decode('utf-8')
                return {
                    "success": False,
                    "message": "No face detected",
                    "processed_frame": f"data:image/jpeg;base64,{frame_b64}"
                }

            for idx, (face_location, face_img) in enumerate(zip(face_locations, processed_faces)):
                # Find best match with lower threshold
                student_id, match_score, student_info = self.face_matcher.find_best_match(
                    face_img, min_threshold=30)
                
                # Check if student already processed
                if student_id and student_info and match_score >= 30:
                    # Draw green box for matched face
                    display_frame = self.image_utils.draw_face_box(
                        display_frame,
                        face_location,
                        name=f"{student_info.get('name', '')} - {match_score:.1f}%",
                        confidence=match_score,
                        color=(0, 255, 0)  # Green box for match
                    )
                    
                    if student_id not in self.processed_students:
                        # Create/Update Excel attendance record
                        excel_file = create_or_update_attendance_excel(student_info, match_score)
                        
                        # Mark student as processed
                        self.processed_students.add(student_id)
                        
                        # Mark attendance in database
                        attendance_data = {
                            'student_id': student_id,
                            'date': datetime.now().date(),
                            'entry_time': datetime.now().time(),
                            'match_confidence': match_score
                        }
                        self.attendance_db.mark_attendance(attendance_data)
                        
                        # Convert processed frame to base64
                        _, buffer = cv2.imencode('.jpg', display_frame)
                        frame_b64 = base64.b64encode(buffer).decode('utf-8')
                        
                        return {
                            "success": True,
                            "name": student_info.get('name', ''),
                            "confidence": match_score,
                            "message": "Attendance marked successfully",
                            "processed_frame": f"data:image/jpeg;base64,{frame_b64}"
                        }
                else:
                    # Draw red box for unmatched face
                    display_frame = self.image_utils.draw_face_box(
                        display_frame,
                        face_location,
                        name="Unknown",
                        confidence=match_score if match_score else 0.0,
                        color=(0, 0, 255)  # Red box for no match
                    )
            
            # Convert final frame to base64
            _, buffer = cv2.imencode('.jpg', display_frame)
            frame_b64 = base64.b64encode(buffer).decode('utf-8')
            
            return {
                "success": False,
                "message": "No match found",
                "processed_frame": f"data:image/jpeg;base64,{frame_b64}"
            }
            
        except Exception as e:
            self.logger.error(f"Error processing frame: {str(e)}")
            return {
                "success": False,
                "message": f"Error processing frame: {str(e)}"
            }

    def run(self):
        """
        Run the attendance system using camera input.
        """
        try:
            cap = cv2.VideoCapture(self.camera_id)
            cap.set(cv2.CAP_PROP_FRAME_WIDTH, self.frame_width)
            cap.set(cv2.CAP_PROP_FRAME_HEIGHT, self.frame_height)
            
            self.logger.info("Starting attendance system")
            
            while True:
                ret, frame = cap.read()
                if not ret:
                    break
                
                # Process frame
                processed_frame = self.process_frame(frame)
                
                # Display
                cv2.imshow(settings.UI['window_title'], processed_frame)
                
                # Break on 'q' or if attendance is recorded
                if cv2.waitKey(1) & 0xFF == ord('q') or len(self.processed_students) > 0:
                    break
                    
            cap.release()
            cv2.destroyAllWindows()
            
        except Exception as e:
            self.logger.error(f"Error in main loop: {str(e)}")
        finally:
            self.cleanup()

    def cleanup(self):
        """Perform system cleanup operations."""
        try:
            self.logger.info("System shutdown complete")
        except Exception as e:
            self.logger.error(f"Error during cleanup: {str(e)}")

# Create global instance of AttendanceSystem
attendance_system = AttendanceSystem()

@app.websocket("/ws/scan")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    try:
        while True:
            # Receive the base64 image from frontend
            data = await websocket.receive_text()
            
            try:
                # Decode base64 image
                encoded_data = data.split(',')[1]
                nparr = np.frombuffer(base64.b64decode(encoded_data), np.uint8)
                frame = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

                # Process frame
                result = await attendance_system.process_frame_websocket(frame)
                
                # Send result back to client
                await websocket.send_json(result)

            except Exception as e:
                print(f"Error processing frame: {str(e)}")
                await websocket.send_json({
                    "success": False,
                    "message": "Error processing image"
                })

    except Exception as e:
        print(f"WebSocket error: {str(e)}")

@app.get("/api/students/count")
async def get_student_count():
    """Get total number of registered students."""
    try:
        students = attendance_system.student_db.get_all_students()
        return {"count": len(students)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/attendance/today")
async def get_today_attendance():
    """Get today's attendance statistics."""
    try:
        today = datetime.now().date()
        all_students = attendance_system.student_db.get_all_students()
        attendance_records = attendance_system.attendance_db.get_daily_attendance("all", today)
        
        return {
            "total_students": len(all_students),
            "present": len(attendance_system.processed_students),
            "absent": len(all_students) - len(attendance_system.processed_students),
            "records": attendance_records
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/system/status")
async def get_system_status():
    """Get system status and configuration."""
    try:
        return {
            "status": "running",
            "camera": {
                "id": settings.CAMERA_ID,
                "width": settings.FRAME_WIDTH,
                "height": settings.FRAME_HEIGHT,
                "fps": settings.FPS
            },
            "face_recognition": {
                "model": settings.FACE_DETECTION_MODEL,
                "threshold": settings.FACE_MATCH_THRESHOLD,
                "min_face_size": settings.MIN_FACE_SIZE
            },
            "attendance_rules": {
                "late_threshold": settings.LATE_THRESHOLD,
                "half_day_threshold": settings.HALF_DAY_THRESHOLD,
                "minimum_attendance": settings.MINIMUM_ATTENDANCE
            }
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/attendance/summary")
async def get_attendance_summary():
    """Get a summary of attendance statistics."""
    try:
        today = datetime.now().date()
        all_students = attendance_system.student_db.get_all_students()
        
        # Get attendance statistics
        stats = attendance_system.attendance_db.get_attendance_statistics(
            class_name="all",
            start_date=today,
            end_date=today
        )
        
        return {
            "date": today.isoformat(),
            "total_students": len(all_students),
            "attendance_percentage": stats.get('attendance_percentage', 0),
            "present_count": stats.get('present_count', 0),
            "absent_count": stats.get('absent_count', 0),
            "late_count": stats.get('late_count', 0),
            "half_day_count": stats.get('half_day_count', 0)
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.websocket("/ws/monitor")
async def monitor_endpoint(websocket: WebSocket):
    """WebSocket endpoint for real-time monitoring."""
    await websocket.accept()
    try:
        while True:
            # Send system status every few seconds
            status = {
                "time": datetime.now().isoformat(),
                "active_users": len(attendance_system.processed_students),
                "system_status": "active",
                "last_detection": None
            }
            await websocket.send_json(status)
            await asyncio.sleep(5)  # Update every 5 seconds
            
    except Exception as e:
        print(f"Monitor WebSocket error: {str(e)}")

def run_standalone():
    """Run the system in standalone mode with OpenCV window."""
    # Add students to database
    add_multiple_students()
    
    # Run the attendance system
    system = AttendanceSystem()
    system.run()

def run_server():
    """Run the system as a FastAPI server."""
    # Add students to database
    add_multiple_students()
    
    # Start the FastAPI server
    uvicorn.run(app, host="0.0.0.0", port=8000)

if __name__ == "__main__":
    # Check command line arguments for mode
    if len(sys.argv) > 1 and sys.argv[1] == "--server":
        print("Starting in server mode...")
        run_server()
    else:
        print("Starting in standalone mode...")
        run_standalone()