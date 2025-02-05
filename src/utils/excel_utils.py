import pandas as pd
from typing import Dict, List, Optional, Union
import os
from datetime import datetime, date
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelManager:
    def __init__(self, output_dir: str = "data/attendance", log_dir: str = "logs"):
        self.output_dir = output_dir
        self.setup_logging(log_dir)
        os.makedirs(output_dir, exist_ok=True)

    def setup_logging(self, log_dir: str) -> None:
        os.makedirs(log_dir, exist_ok=True)
        log_file = os.path.join(log_dir, f'excel_utils_{datetime.now().strftime("%Y%m%d")}.log')
        logging.basicConfig(
            filename=log_file,
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s'
        )
        self.logger = logging.getLogger(__name__)

    def create_daily_report(self, 
                          attendance_data: List[Dict],
                          class_name: str,
                          report_date: date) -> str:
        try:
            df = pd.DataFrame(attendance_data)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = f"Attendance_{report_date.strftime('%Y%m%d')}"
            
            # Add headers
            headers = [
                'Roll No', 'Name', 'Entry Time', 'Exit Time', 
                'Status', 'Match Confidence', 'Remarks'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row_idx, record in enumerate(attendance_data, 2):
                ws.cell(row=row_idx, column=1).value = record['roll_number']
                ws.cell(row=row_idx, column=2).value = record['name']
                ws.cell(row=row_idx, column=3).value = record.get('entry_time', '')
                ws.cell(row=row_idx, column=4).value = record.get('exit_time', '')
                ws.cell(row=row_idx, column=5).value = record['status']
                ws.cell(row=row_idx, column=6).value = f"{record.get('match_confidence', 0):.2f}%"
                ws.cell(row=row_idx, column=7).value = record.get('remarks', '')
                
                # Color coding for status
                status_colors = {
                    'present': '90EE90',  # Light green
                    'absent': 'FFB6C1',   # Light red
                    'late': 'FFD700',     # Gold
                    'half-day': 'F0E68C'  # Khaki
                }
                ws.cell(row=row_idx, column=5).fill = PatternFill(
                    start_color=status_colors.get(record['status'], 'FFFFFF'),
                    end_color=status_colors.get(record['status'], 'FFFFFF'),
                    fill_type='solid'
                )
            
            # Add summary
            summary_row = len(attendance_data) + 3
            ws.cell(row=summary_row, column=1).value = "Summary"
            ws.cell(row=summary_row, column=1).font = Font(bold=True)
            
            status_counts = pd.Series([r['status'] for r in attendance_data]).value_counts()
            for idx, (status, count) in enumerate(status_counts.items(), 1):
                ws.cell(row=summary_row + idx, column=1).value = f"Total {status.title()}"
                ws.cell(row=summary_row + idx, column=2).value = count
            
            # Format columns
            for column in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(column)].width = 15
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1, max_row=len(attendance_data) + 1):
                for cell in row:
                    cell.border = thin_border
            
            # Save file
            filename = f"attendance_{class_name}_{report_date.strftime('%Y%m%d')}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            self.logger.info(f"Created daily report: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error creating daily report: {str(e)}")
            return ""

    def create_monthly_report(self,
                            attendance_data: pd.DataFrame,
                            class_name: str,
                            month: int,
                            year: int) -> str:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = f"Monthly_Attendance_{year}_{month:02d}"
            
            # Add headers
            headers = ['Roll No', 'Name'] + [str(i) for i in range(1, 32)] + [
                'Total Present', 'Total Absent', 'Total Late', 'Total Half-Day', 'Attendance %'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Add data
            for row_idx, (idx, row) in enumerate(attendance_data.iterrows(), 2):
                ws.cell(row=row_idx, column=1).value = idx[0]  # Roll number
                ws.cell(row=row_idx, column=2).value = idx[1]  # Name
                
                # Daily attendance
                for day in range(1, 32):
                    if f"{year}-{month:02d}-{day:02d}" in row.index:
                        status = row[f"{year}-{month:02d}-{day:02d}"]
                        cell = ws.cell(row=row_idx, column=day + 2)
                        cell.value = status
                        
                        # Color coding
                        status_colors = {
                            'present': '90EE90',
                            'absent': 'FFB6C1',
                            'late': 'FFD700',
                            'half-day': 'F0E68C'
                        }
                        cell.fill = PatternFill(
                            start_color=status_colors.get(status, 'FFFFFF'),
                            end_color=status_colors.get(status, 'FFFFFF'),
                            fill_type='solid'
                        )
                
                # Summary columns
                summary_start = 34  # Column after days
                ws.cell(row=row_idx, column=summary_start).value = row['Total Present']
                ws.cell(row=row_idx, column=summary_start + 1).value = row['Total Absent']
                ws.cell(row=row_idx, column=summary_start + 2).value = row['Total Late']
                ws.cell(row=row_idx, column=summary_start + 3).value = row['Total Half-Day']
                
                # Calculate attendance percentage
                total_days = row['Total Present'] + row['Total Absent'] + \
                           row['Total Late'] + row['Total Half-Day']
                attendance_pct = (row['Total Present'] + 0.5 * row['Total Half-Day']) / total_days * 100
                ws.cell(row=row_idx, column=summary_start + 4).value = f"{attendance_pct:.1f}%"
            
            # Format columns
            for column in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(column)].width = 12
            
            # Add borders
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            for row in ws.iter_rows(min_row=1, max_row=len(attendance_data) + 1):
                for cell in row:
                    cell.border = thin_border
            
            # Save file
            filename = f"monthly_attendance_{class_name}_{year}_{month:02d}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            self.logger.info(f"Created monthly report: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error creating monthly report: {str(e)}")
            return ""

    def export_attendance_statistics(self,
                                  stats_data: Dict,
                                  class_name: str,
                                  start_date: date,
                                  end_date: date) -> str:
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance_Statistics"
            
            # Add title
            ws['A1'] = f"Attendance Statistics for {class_name}"
            ws['A2'] = f"Period: {start_date} to {end_date}"
            ws['A1'].font = Font(bold=True, size=14)
            ws['A2'].font = Font(italic=True)
            
            # Add statistics
            stats_row = 4
            stats_mapping = {
                'total_records': 'Total Records',
                'present_count': 'Total Present',
                'absent_count': 'Total Absent',
                'late_count': 'Total Late',
                'half_day_count': 'Total Half Days',
                'attendance_percentage': 'Attendance Percentage'
            }
            
            for key, label in stats_mapping.items():
                ws.cell(row=stats_row, column=1).value = label
                value = stats_data.get(key, 0)
                if key == 'attendance_percentage':
                    ws.cell(row=stats_row, column=2).value = f"{value:.1f}%"
                else:
                    ws.cell(row=stats_row, column=2).value = value
                stats_row += 1
            
            # Format worksheet
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 15
            
            # Add borders and styling
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in range(4, stats_row):
                ws.cell(row=row, column=1).border = border
                ws.cell(row=row, column=2).border = border
                ws.cell(row=row, column=1).font = Font(bold=True)
                ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')
            
            # Save file
            filename = f"attendance_stats_{class_name}_{start_date}_{end_date}.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            wb.save(filepath)
            
            self.logger.info(f"Created statistics report: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error creating statistics report: {str(e)}")
            return ""